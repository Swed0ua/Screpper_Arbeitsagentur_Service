"""Excel file processor for reading and writing company data."""

import csv
import os
import math
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment


class ExcelProcessor:
    """Process Excel/CSV files with company data."""
    
    def __init__(self, file_path: str):
        """
        Initialize Excel processor.
        
        Args:
            file_path: Path to Excel or CSV file
        """
        self.file_path = file_path
        self.df = None
    
    async def load_file(self):
        """Load Excel or CSV file into DataFrame."""
        file_extension = os.path.splitext(self.file_path)[1].lower()
        
        if file_extension == '.csv':
            self.df = pd.read_csv(self.file_path)
        else:
            self.df = pd.read_excel(self.file_path)
        
        return self.df
    
    def get_companies_data(self) -> list:
        """
        Extract company data from DataFrame.
        
        Returns:
            List of dictionaries with company information
        """
        if self.df is None:
            raise ValueError("File not loaded. Call load_file() first.")
        
        companies = []
        for _, row in self.df.iterrows():
            company_data = {
                'company_name': self._get_value(row, 'Роботодавець') or self._get_value(row, 'employer_company_name', ''),
                'title': self._get_value(row, 'Назва вакансії') or self._get_value(row, 'title', ''),
                'location': self._get_value(row, 'Місцезнаходження') or self._get_value(row, 'location', ''),
                'type_offer': self._get_value(row, 'Тип зайнятості') or self._get_value(row, 'type_offer', ''),
                'email': self._get_value(row, 'Електронна пошта') or self._get_value(row, 'email', ''),
                'phone': self._get_value(row, 'Мобільний номер') or self._get_value(row, 'phone', ''),
                'link': self._get_value(row, 'Посилання на вакансію') or self._get_value(row, 'link', ''),
                'job_title': self._get_value(row, 'Назва вакансії') or self._get_value(row, 'title', ''),
                'address': self._get_value(row, 'Поштова адреса') or self._get_value(row, 'address', ''),
            }
            companies.append(company_data)
        
        return companies
    
    def _get_value(self, row, key: str, default: str = '') -> str:
        """Helper to safely get value from row."""
        if key in row.index:
            value = row[key]
            return str(value) if pd.notna(value) else default
        return default
    
    def add_company_research_column(self, company_researches: list):
        """
        Add company research column to DataFrame.
        
        Args:
            company_researches: List of company research strings
        """
        if self.df is None:
            raise ValueError("File not loaded. Call load_file() first.")
        
        self.df['company_research'] = company_researches
    
    def add_suitability_columns(self, suitability_results: list):
        """
        Add suitability columns to DataFrame.
        
        Args:
            suitability_results: List of dicts with 'is_suitable', 'industry', 'rejection_reason'
        """
        if self.df is None:
            raise ValueError("File not loaded. Call load_file() first.")
        
        # Extract values from suitability results
        is_suitable_list = []
        industry_list = []
        rejection_reason_list = []
        
        for result in suitability_results:
            is_suitable_list.append(result.get('is_suitable', False))
            industry_list.append(result.get('industry', 'Unknown'))
            rejection_reason_list.append(result.get('rejection_reason', ''))
        
        # Add columns to DataFrame
        self.df['is_suitable'] = is_suitable_list
        self.df['industry'] = industry_list
        self.df['rejection_reason'] = rejection_reason_list
    
    def add_email_column(self, email_contents: list):
        """
        Add email content column to DataFrame.
        
        Args:
            email_contents: List of email content strings
        """
        if self.df is None:
            raise ValueError("File not loaded. Call load_file() first.")
        
        self.df['email_content'] = email_contents
    
    def _split_long_content(self, content: str, chunk_size: int = 32000) -> list:
        """
        Розбиває довгий текст на частини для Excel.
        
        Args:
            content: Текст для розбиття
            chunk_size: Розмір частини (за замовчуванням 32000, щоб був запас від ліміту Excel 32767)
        
        Returns:
            Список частин тексту
        """
        if not content:
            return [""]
        
        content_str = str(content)
        parts = []
        
        # Розбиваємо на частини по chunk_size символів
        for i in range(0, len(content_str), chunk_size):
            parts.append(content_str[i:i + chunk_size])
        
        return parts
    
    def _prepare_dataframe_for_excel(self):
        """
        Підготовка DataFrame перед збереженням: розбиває довгі HTML колонки на частини.
        Excel має ліміт 32767 символів на комірку, тому довгі HTML (>32000) треба розбивати на кілька колонок.
        """
        if self.df is None:
            return
        
        EXCEL_CELL_LIMIT = 32767
        CHUNK_SIZE = 32000  # Розмір частини (з запасом від ліміту)
        
        html_columns = ['email_content', 'company_research']
        
        # Знаходимо максимальну кількість частин для кожної HTML колонки
        max_parts_needed = {}
        
        for col in html_columns:
            if col not in self.df.columns:
                continue
            
            max_parts = 1
            for idx, value in self.df[col].items():
                if pd.notna(value):
                    value_str = str(value)
                    if len(value_str) > CHUNK_SIZE:
                        num_parts = math.ceil(len(value_str) / CHUNK_SIZE)
                        max_parts = max(max_parts, num_parts)
            
            if max_parts > 1:
                max_parts_needed[col] = max_parts
                max_content_len = self.df[col].astype(str).str.len().max()
                print(f"Column '{col}' will be split into {max_parts} parts (max content: {max_content_len} chars)")
        
        # Створюємо нові колонки для частин та заповнюємо їх
        for col, num_parts in max_parts_needed.items():
            # Створюємо нові колонки
            for part_num in range(2, num_parts + 1):
                part_col_name = f"{col}_part{part_num}"
                if part_col_name not in self.df.columns:
                    self.df[part_col_name] = ""
            
            # Розбиваємо контент для кожного рядка
            for idx in self.df.index:
                value = self.df.at[idx, col]
                
                if pd.notna(value):
                    value_str = str(value)
                    
                    if len(value_str) > CHUNK_SIZE:
                        # Розбиваємо на частини
                        parts = self._split_long_content(value_str, CHUNK_SIZE)
                        
                        # Перша частина залишається в оригінальній колонці
                        self.df.at[idx, col] = parts[0] if parts else ""
                        
                        # Додаткові частини додаємо в нові колонки
                        for part_num in range(1, len(parts)):
                            part_col_name = f"{col}_part{part_num + 1}"
                            self.df.at[idx, part_col_name] = parts[part_num] if part_num < len(parts) else ""
                    else:
                        # Якщо не перевищує ліміт, заповнюємо нові колонки порожніми рядками
                        for part_num in range(2, num_parts + 1):
                            part_col_name = f"{col}_part{part_num}"
                            self.df.at[idx, part_col_name] = ""
    
    async def save_file(self, output_path: str = None):
        """
        Save DataFrame to Excel file with formatting.
        Автоматично розбиває довгий HTML на кілька колонок, щоб уникнути обрізання.
        
        Args:
            output_path: Output file path (if None, overwrites original)
        """
        if self.df is None:
            raise ValueError("No data to save.")
        
        save_path = output_path or self.file_path
        file_extension = os.path.splitext(save_path)[1].lower()
        
        if file_extension == '.csv':
            # For CSV, ensure HTML is properly saved (no truncation)
            self.df.to_csv(save_path, index=False, encoding='utf-8', quoting=csv.QUOTE_ALL)
            print(f"File saved: {save_path}")
        else:
            # Для Excel - спочатку підготовка DataFrame (розбиття довгих HTML)
            print("Preparing DataFrame: splitting long HTML content into multiple columns...")
            self._prepare_dataframe_for_excel()
            
            # Використовуємо openpyxl напряму
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            
            # Записати заголовки
            headers = list(self.df.columns)
            ws.append(headers)
            
            EXCEL_CELL_LIMIT = 32767  # Фізичний ліміт Excel
            
            # Записати дані рядок за рядком
            for idx, row in self.df.iterrows():
                row_data = []
                for col in headers:
                    value = row[col]
                    
                    if pd.notna(value):
                        value_str = str(value)
                        
                        # Логування для HTML колонок
                        if 'email_content' in col or 'company_research' in col:
                            print(f"Row {idx + 1}, Column {col}: {len(value_str)} chars")
                        
                        # Перевірка на Excel ліміт (на всяк випадок - має бути вже розбито)
                        if len(value_str) > EXCEL_CELL_LIMIT:
                            print(f"ERROR: Content STILL exceeds Excel limit! Row {idx + 1}, Column {col}: {len(value_str)} chars")
                            print(f"  This should not happen after splitting. Truncating to {EXCEL_CELL_LIMIT} chars.")
                            value_str = value_str[:EXCEL_CELL_LIMIT]
                        
                        row_data.append(value_str)
                    else:
                        row_data.append("")
                
                ws.append(row_data)
            
            # Встановити ширину колонок
            for idx, col_name in enumerate(headers, 1):
                column_letter = get_column_letter(idx)
                
                if 'company_research' in col_name or 'email_content' in col_name:
                    ws.column_dimensions[column_letter].width = 100
                elif idx == 1:
                    ws.column_dimensions[column_letter].width = 30
                else:
                    ws.column_dimensions[column_letter].width = 50
            
            # Встановити форматування для всіх комірок
            for row_num in range(1, ws.max_row + 1):
                for col_num, col_name in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    
                    if cell.value and isinstance(cell.value, str):
                        cell.alignment = Alignment(
                            wrap_text=True,
                            vertical='top',
                            horizontal='left'
                        )
            
            # Зберегти файл
            wb.save(save_path)
            print(f"File saved: {save_path}")

